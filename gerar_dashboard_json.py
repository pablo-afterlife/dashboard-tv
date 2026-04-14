#!/usr/bin/env python3
import json
import math
import os
import sys
from datetime import datetime, date
from openpyxl import load_workbook

INPUT_XLSX = sys.argv[1] if len(sys.argv) > 1 else "Andamento - Projetos.xlsx"
SHEET_NAME = sys.argv[2] if len(sys.argv) > 2 else "filtrada"
OUTPUT_JSON = sys.argv[3] if len(sys.argv) > 3 else "dashboard_tv_data.json"

def clean_value(value):
    if value is None:
        return None
    if isinstance(value, (datetime, date)):
        return value.strftime("%Y-%m-%d")
    if isinstance(value, (int,)):
        return value
    if isinstance(value, float):
        if not math.isfinite(value):
            return None
        if abs(value - round(value)) < 1e-9:
            return int(round(value))
        return value
    if isinstance(value, str):
        value = value.strip()
        if value == "" or value == "-":
            return None
        return value
    return str(value)

def main():
    wb = load_workbook(INPUT_XLSX, data_only=True)
    if SHEET_NAME not in wb.sheetnames:
        raise SystemExit(f"Aba não encontrada: {SHEET_NAME}")

    ws = wb[SHEET_NAME]
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        raise SystemExit("A planilha está vazia.")

    headers = [str(h).strip() if h is not None else f"COL_{i+1}" for i, h in enumerate(rows[0])]
    records = []
    for row in rows[1:]:
        if not any(cell is not None and str(cell).strip() != "" for cell in row):
            continue
        item = {}
        for header, value in zip(headers, row):
            item[header] = clean_value(value)
        records.append(item)

    payload = {
        "source_file": os.path.basename(INPUT_XLSX),
        "source_sheet": SHEET_NAME,
        "generated_at": datetime.now().astimezone().isoformat(timespec="seconds"),
        "row_count": len(records),
        "records": records,
    }

    with open(OUTPUT_JSON, "w", encoding="utf-8") as f:
        json.dump(payload, f, ensure_ascii=False, indent=2)

    print(f"JSON gerado com sucesso: {OUTPUT_JSON}")
    print(f"Registros exportados: {len(records)}")

if __name__ == "__main__":
    main()
